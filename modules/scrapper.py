
import requests
from bs4 import BeautifulSoup
from utils import errhandler, syshandler
from helpers import clean_citation, clean_citation_text, get_court_type

from pathlib import Path
from typing import Optional, List, Dict, Any
import urllib.parse
from difflib import SequenceMatcher
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
import time
import re
import traceback


class Scrapper:
    """
    KRA iLaw scrapper with fixed authentication and improved matching.
    """

    def __init__(
        self,
        session=None,
        auth_url: str = "",
        url: str = "",
        keyword: str = "",
        data: list | None = None,
        username: str = "",
        password: str = "",
    ):
        self.session      = session or requests.Session()
        self.auth_url     = auth_url or "https://ilaw.kra.go.ke/ilaw/users/login"
        self.url          = url or "https://ilaw.kra.go.ke/ilaw/search/universal/1?keyword="
        self.data         = data
        self.username     = username
        self.password     = password
        self.authenticated = False
        self.results      = []

        self.session.headers.update({
            "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate",
            "Connection":      "keep-alive",
            "Upgrade-Insecure-Requests": "1",
        })

    # ─────────────────────────────────────────────────────────────────────────
    # Authentication
    # ─────────────────────────────────────────────────────────────────────────

    def authenticator(self) -> bool:
        print(f"\n🔐 Authenticating to KRA iLaw with username: {self.username}")

        if self._try_auth_method_1():
            return True
        if self._try_auth_method_2():
            return True
        if self._try_auth_method_3():
            return True

        print("❌ All authentication methods failed")
        return False

    def _detect_login_fields(self, soup) -> dict:
        """
        Detect the actual username and password field names from the login form.
        Skips hidden fields and fields named 'redirect_to' / token fields.
        Returns: {'username_field': '...', 'password_field': '...', 'hidden': {...}}
        """
        # Username/password type hints
        USERNAME_HINTS = ['username', 'user', 'email', 'login', 'userid', 'user_name', 'usr', 'account']
        PASSWORD_HINTS = ['password', 'pass', 'passwd', 'pwd']

        # Fields to explicitly skip as username candidates
        SKIP_NAMES = {'redirect_to', 'csrf_token', 'csrfmiddlewaretoken', '_token',
                      'authenticity_token', 'token', 'nonce', 'submit', 'login_btn'}

        form = None
        # Pick the form that contains a password input
        for f in soup.find_all('form'):
            if f.find('input', {'type': 'password'}):
                form = f
                break
        if not form:
            form = soup  # search whole page as fallback

        all_inputs = form.find_all('input') if form else soup.find_all('input')

        username_field = None
        password_field = None
        hidden_fields  = {}

        # Collect hidden fields (CSRF tokens etc.)
        for inp in all_inputs:
            t    = (inp.get('type') or 'text').lower()
            name = inp.get('name', '')
            if t == 'hidden' and name:
                hidden_fields[name] = inp.get('value', '')

        # Find password field first (unambiguous)
        for inp in all_inputs:
            t = (inp.get('type') or 'text').lower()
            if t == 'password':
                password_field = inp.get('name', 'password')
                break

        # Find username field — must be a visible text/email input
        for inp in all_inputs:
            t    = (inp.get('type') or 'text').lower()
            name = (inp.get('name') or '').lower()
            iid  = (inp.get('id') or '').lower()
            ph   = (inp.get('placeholder') or '').lower()

            if t in ('hidden', 'password', 'submit', 'button', 'checkbox', 'radio'):
                continue
            if name in SKIP_NAMES or iid in SKIP_NAMES:
                continue

            # Check against hints
            for hint in USERNAME_HINTS:
                if hint in name or hint in iid or hint in ph:
                    username_field = inp.get('name')
                    break
            if username_field:
                break

        # If no hint match, fall back to first non-hidden, non-password visible input
        if not username_field:
            for inp in all_inputs:
                t    = (inp.get('type') or 'text').lower()
                name = (inp.get('name') or '').lower()
                if t in ('text', 'email') and name not in SKIP_NAMES:
                    username_field = inp.get('name')
                    break

        result = {
            'username_field': username_field or 'username',
            'password_field': password_field or 'password',
            'hidden':         hidden_fields,
        }
        print(f"✅ Login form detected: user='{result['username_field']}', pass='{result['password_field']}'")
        return result

    def _is_logged_in(self, response) -> bool:
        """Returns True if the response indicates a successful login."""
        # Must NOT still be on the login page
        if self.auth_url in response.url or 'login' in response.url.lower():
            # Even if on login URL, check if page contains dashboard content
            pass

        # Strong positive signals
        positive = ['dashboard', 'welcome', 'logout', 'log out', 'my account',
                    'search cases', 'ilaw home', 'ilaw/search']
        text_low = response.text.lower()
        for sig in positive:
            if sig in text_low:
                return True

        # Redirected away from login URL = success
        if response.url and self.auth_url not in response.url:
            if 'login' not in response.url.lower():
                return True

        # No password field on resulting page = success
        soup = BeautifulSoup(response.text, 'html.parser')
        pw_inputs = soup.find_all('input', {'type': 'password'})
        if not pw_inputs and response.status_code == 200:
            return True

        return False

    def _try_auth_method_1(self) -> bool:
        """Primary: detect fields from live page, POST with correct field names."""
        try:
            print("🔄 Trying Method 1: Smart field detection...")
            login_page = self.session.get(self.auth_url, timeout=15)
            print(f"📥 Login page status: {login_page.status_code}")

            soup   = BeautifulSoup(login_page.text, 'html.parser')
            fields = self._detect_login_fields(soup)

            payload = {
                fields['username_field']: self.username,
                fields['password_field']: self.password,
            }
            # Add hidden fields (CSRF tokens etc.)
            payload.update(fields['hidden'])

            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Origin':       'https://ilaw.kra.go.ke',
                'Referer':      self.auth_url,
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'same-origin',
                'Sec-Fetch-User': '?1',
            }

            response = self.session.post(
                self.auth_url, data=payload, headers=headers,
                allow_redirects=True, timeout=30
            )
            print(f"📥 Auth response status: {response.status_code}")
            print(f"📥 Final URL: {response.url}")

            if self._is_logged_in(response):
                print("✅ Method 1 successful")
                self.authenticated = True
                return True

            print("❌ Method 1 failed")
            return False
        except Exception as e:
            print(f"❌ Method 1 error: {e}")
            return False

    def _try_auth_method_2(self) -> bool:
        """Fallback: fresh session, visit homepage first, then POST."""
        try:
            print("🔄 Trying Method 2: Fresh session with homepage priming...")
            self.session = requests.Session()
            self.session.headers.update({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            })

            # Visit homepage to pick up cookies
            self.session.get("https://ilaw.kra.go.ke", timeout=10)
            login_page = self.session.get(self.auth_url, timeout=15)
            soup   = BeautifulSoup(login_page.text, 'html.parser')
            fields = self._detect_login_fields(soup)

            payload = {
                fields['username_field']: self.username,
                fields['password_field']: self.password,
            }
            payload.update(fields['hidden'])

            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Origin':       'https://ilaw.kra.go.ke',
                'Referer':      self.auth_url,
            }

            response = self.session.post(
                self.auth_url, data=payload, headers=headers,
                allow_redirects=True, timeout=30
            )

            if self._is_logged_in(response):
                print("✅ Method 2 successful")
                self.authenticated = True
                return True

            print("❌ Method 2 failed")
            return False
        except Exception as e:
            print(f"❌ Method 2 error: {e}")
            return False

    def _try_auth_method_3(self) -> bool:
        """Last resort: full Chromium-like headers, hardcoded field names."""
        try:
            print("🔄 Trying Method 3: Full browser headers...")

            full_headers = {
                'User-Agent':      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
                'Accept':          'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Content-Type':    'application/x-www-form-urlencoded',
                'Origin':          'https://ilaw.kra.go.ke',
                'Referer':         self.auth_url,
                'Sec-Ch-Ua':       '"Chromium";v="122", "Not(A:Brand";v="24"',
                'Sec-Ch-Ua-Mobile': '?0',
                'Sec-Ch-Ua-Platform': '"Windows"',
                'Sec-Fetch-Dest':  'document',
                'Sec-Fetch-Mode':  'navigate',
                'Sec-Fetch-Site':  'same-origin',
                'Sec-Fetch-User':  '?1',
                'Upgrade-Insecure-Requests': '1',
            }
            self.session.headers.update(full_headers)

            login_page = self.session.get(self.auth_url, timeout=15)
            soup   = BeautifulSoup(login_page.text, 'html.parser')
            fields = self._detect_login_fields(soup)

            # Also try hardcoded fallback names
            for uname in [fields['username_field'], 'username', 'user', 'email']:
                for pname in [fields['password_field'], 'password', 'pass']:
                    payload = {uname: self.username, pname: self.password}
                    payload.update(fields['hidden'])
                    response = self.session.post(
                        self.auth_url, data=payload,
                        allow_redirects=True, timeout=30
                    )
                    if self._is_logged_in(response):
                        print(f"✅ Method 3 successful (user='{uname}', pass='{pname}')")
                        self.authenticated = True
                        return True

            print("❌ Method 3 failed")
            return False
        except Exception as e:
            print(f"❌ Method 3 error: {e}")
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Extractor
    # ─────────────────────────────────────────────────────────────────────────

    def extractor(self) -> Optional[List[Dict[str, Any]]]:
        if not self.authenticated:
            print("❌ Not authenticated. Please login first.")
            return None
        if not self.data:
            print("⚠️ No data to process")
            return []

        final_results = []
        print(f"⌛ Starting extraction for {len(self.data)} records...")

        ajax_headers = {
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type":     "application/x-www-form-urlencoded",
            "Referer":          "https://ilaw.kra.go.ke/ilaw/search/universal",
        }
        payload = {"dataType": "litigation_data"}

        for i, item in enumerate(self.data):
            if i > 0 and i % 10 == 0:
                time.sleep(2)   # polite rate limiting

            keyword    = item['keyword']
            safe_kw    = urllib.parse.quote(keyword)
            query_url  = f"{self.url}{safe_kw}"

            print(f"\n{'='*50}")
            print(f"🔎 Record {i+1}/{len(self.data)}")
            print(f"📋 Case:    {item.get('case_number', 'N/A')}")
            print(f"🔑 Keyword: {keyword}")

            try:
                response = self.session.post(
                    query_url, data=payload, headers=ajax_headers, timeout=30
                )

                if response.status_code != 200:
                    print(f"⚠️ HTTP {response.status_code} — trying re-auth...")
                    if self.authenticator():
                        response = self.session.post(
                            query_url, data=payload, headers=ajax_headers, timeout=30
                        )
                    if response.status_code != 200:
                        final_results.append(self._empty_result(item))
                        continue

                try:
                    json_data = response.json()
                except json.JSONDecodeError:
                    # Probably redirected to login page
                    soup = BeautifulSoup(response.text, 'html.parser')
                    if soup.find('input', {'type': 'password'}):
                        print("⚠️ Session expired — re-authenticating...")
                        if self.authenticator():
                            response = self.session.post(
                                query_url, data=payload, headers=ajax_headers, timeout=30
                            )
                            try:
                                json_data = response.json()
                            except Exception:
                                final_results.append(self._empty_result(item))
                                continue
                        else:
                            break
                    else:
                        final_results.append(self._empty_result(item))
                        continue

                html_string = json_data.get('html', '')
                if not html_string:
                    print("⚠️ No HTML in response — no results for this keyword")
                    final_results.append(self._empty_result(item))
                    continue

                found_matches = self._parse_results(html_string)

                result_entry = {
                    "excel_row":      item.get('excel_row'),
                    "original_case":  item['case_number'],
                    "case_name":      item['citation'],
                    "search_keyword": keyword,
                    "matches_found":  len(found_matches),
                    "matches":        found_matches,
                }
                final_results.append(result_entry)
                print(f"📊 Found {len(found_matches)} matches")

            except requests.exceptions.Timeout:
                print(f"⏰ Timeout for record {i+1} — skipping")
                final_results.append(self._empty_result(item))
            except Exception as e:
                errhandler(f"Error processing record {i+1}: {e}", log="extractor", path="scrapper")
                print(f"❌ Error: {e}")
                traceback.print_exc()
                final_results.append(self._empty_result(item))

        print(f"\n{'='*50}")
        print(f"✅ Extraction complete. Processed {len(final_results)} records")
        return final_results

    def _empty_result(self, item: dict) -> dict:
        return {
            "excel_row":      item.get('excel_row'),
            "original_case":  item.get('case_number', ''),
            "case_name":      item.get('citation', ''),
            "search_keyword": item.get('keyword', ''),
            "matches_found":  0,
            "matches":        [],
        }

    def _parse_results(self, html_string: str) -> list:
        """Parse KRA iLaw HTML result table into a list of match dicts."""
        soup    = BeautifulSoup(html_string, 'html.parser')
        matches = []
        for row in soup.find_all('tr'):
            cols = row.find_all('td')
            if len(cols) < 3:
                continue

            # Citation — prefer tooltip title if present
            span = cols[1].find('span', class_='tooltipTable') if len(cols) > 1 else None
            if span and span.get('tooltiptitle'):
                citation = span.get('tooltiptitle')
            else:
                citation = cols[1].get_text(strip=True) if len(cols) > 1 else ''

            internal_ref = cols[2].get_text(strip=True) if len(cols) > 2 else ''
            assignee     = cols[3].get_text(strip=True) if len(cols) > 3 else ''

            entry = {
                "kra_citation": self._clean_text(citation),
                "kra_ref":      self._clean_text(internal_ref).upper(),
                "kra_assignee": self._clean_text(assignee).upper(),
            }
            if entry["kra_citation"] or entry["kra_ref"]:
                matches.append(entry)
                print(f"  ✅ Match: {entry['kra_citation'][:60]}")
        return matches

    # ─────────────────────────────────────────────────────────────────────────
    # Comparator
    # ─────────────────────────────────────────────────────────────────────────

    def comparator(self, extracted_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Score each extracted result against the original file data.
        Uses case-number E-code matching, party name fuzzy matching, and
        keyword (E017 of 2026) direct matching as primary signal.
        """
        if not extracted_data:
            print("⚠️ No extracted data to compare")
            return []

        print(f"\n⚖️  Comparing {len(extracted_data)} records...")
        reconciled_data = []

        for item in extracted_data:
            sheet_citation = str(item.get('case_name', '')).upper()
            sheet_case     = str(item.get('original_case', '')).upper()
            search_keyword = str(item.get('search_keyword', '')).upper()

            print(f"\n{'='*50}")
            print(f"🏁 [{sheet_case}] {sheet_citation[:80]}")

            sheet_tokens = clean_citation(sheet_citation)
            sheet_string = clean_citation_text(sheet_citation)
            sheet_court  = get_court_type(sheet_case)

            matches    = item.get('matches', [])
            best_match = {}
            confidence = 0.0
            status     = "NOT FOUND"

            if matches:
                best_ratio = 0.0

                for match in matches:
                    kra_citation = str(match.get('kra_citation', '')).upper()
                    kra_ref      = str(match.get('kra_ref', '')).upper()
                    kra_court    = get_court_type(kra_citation)

                    # Skip cross-court matches (only when both courts are identifiable)
                    if sheet_court != 'NA' and kra_court != 'NA' and sheet_court != kra_court:
                        continue

                    scores = self._calculate_similarity_scores(
                        sheet_tokens, sheet_string, sheet_case,
                        kra_citation, kra_ref, search_keyword
                    )

                    # Weighted final score
                    weighted = (
                        scores['token_ratio']   * 0.30 +
                        scores['string_ratio']  * 0.30 +
                        scores['ref_ratio']     * 0.25 +
                        scores['keyword_ratio'] * 0.15
                    )
                    final = max(
                        weighted,
                        scores['ref_ratio'],      # exact case-number match overrides
                        scores['keyword_ratio'],  # exact keyword hit overrides
                    )

                    if final > best_ratio:
                        best_ratio = final
                        best_match = match

                confidence = round(best_ratio, 2)

                if confidence >= 80:
                    status = "VERIFIED MATCH"
                elif confidence >= 60:
                    status = "REVIEW REQUIRED"
                elif confidence >= 30:
                    status = "MISMATCH"
                else:
                    status = "NOT FOUND"

                print(f"📊 Best match: {status} ({confidence}%)")

            reconciled_data.append({
                'excel_row':               item.get('excel_row'),
                'original_case':           item.get('original_case', ''),
                'case_name':               item.get('case_name', ''),
                'status':                  status,
                'confidence_score':        f"{confidence}%",
                'confidence_raw':          confidence,
                'best_match_kra_ref':      best_match.get('kra_ref', 'N/A') if best_match else 'N/A',
                'best_match_kra_citation': best_match.get('kra_citation', 'N/A') if best_match else 'N/A',
                'best_match_kra_assignee': best_match.get('kra_assignee', 'N/A') if best_match else 'N/A',
                'matches_found':           len(matches),
            })

        # Summary
        counts = {}
        for r in reconciled_data:
            counts[r['status']] = counts.get(r['status'], 0) + 1
        total = len(reconciled_data)
        print(f"\n{'='*50}")
        print("✅ Comparison complete. Summary:")
        for s, c in counts.items():
            print(f"   {s}: {c}/{total} ({round(c/total*100,1)}%)")
        return reconciled_data

    def _calculate_similarity_scores(
        self,
        sheet_tokens, sheet_string, sheet_case,
        kra_citation, kra_ref,
        search_keyword=""
    ) -> dict:
        scores = {}

        # Token overlap (party names, court names)
        kra_tokens = clean_citation(kra_citation)
        if sheet_tokens and kra_tokens:
            inter = sheet_tokens.intersection(kra_tokens)
            union = sheet_tokens.union(kra_tokens)
            scores['token_ratio'] = (len(inter) / len(union)) * 100 if union else 0
        else:
            scores['token_ratio'] = 0

        # Fuzzy string match
        kra_string = clean_citation_text(kra_citation)
        scores['string_ratio'] = SequenceMatcher(None, sheet_string, kra_string).ratio() * 100

        # Case reference number match
        # Extracts E-code (e.g. E017) and year from both sides and compares
        scores['ref_ratio'] = 0
        e_sheet = re.search(r'E(\d+)', sheet_case, re.IGNORECASE)
        e_kra   = re.search(r'E(\d+)', kra_citation + ' ' + kra_ref, re.IGNORECASE)
        yr_sheet = re.search(r'\b(20\d{2}|19\d{2})\b', sheet_case)
        yr_kra   = re.search(r'\b(20\d{2}|19\d{2})\b', kra_citation + ' ' + kra_ref)

        if e_sheet and e_kra and e_sheet.group(1) == e_kra.group(1):
            scores['ref_ratio'] = 80   # E-number matches
            if yr_sheet and yr_kra and yr_sheet.group(0) == yr_kra.group(0):
                scores['ref_ratio'] = 100  # E-number + year both match
        elif sheet_case and kra_ref:
            # Fallback: last number in case vs last number in ref
            sn = re.findall(r'\d+', sheet_case)
            kn = re.findall(r'\d+', kra_ref)
            if sn and kn and sn[-1] == kn[-1]:
                scores['ref_ratio'] = 60

        # Keyword match — does the kra_citation contain our search keyword?
        # e.g. search_keyword="E017 of 2026" and kra_citation contains "E017" and "2026"
        scores['keyword_ratio'] = 0
        if search_keyword:
            kw_e  = re.search(r'E(\d+)', search_keyword, re.IGNORECASE)
            kw_yr = re.search(r'\b(20\d{2}|19\d{2})\b', search_keyword)
            kra_text = (kra_citation + ' ' + kra_ref).upper()
            if kw_e and kw_yr:
                has_e  = bool(re.search(rf'E0*{kw_e.group(1)}\b', kra_text, re.IGNORECASE))
                has_yr = kw_yr.group(0) in kra_text
                if has_e and has_yr:
                    scores['keyword_ratio'] = 100
                elif has_e or has_yr:
                    scores['keyword_ratio'] = 50

        return scores

    def _clean_text(self, text) -> str:
        if not text:
            return ""
        text = ' '.join(str(text).split())
        return text.encode('ascii', 'ignore').decode().strip()

    # ─────────────────────────────────────────────────────────────────────────
    # Reporter
    # ─────────────────────────────────────────────────────────────────────────

    def report(self, data: list | None = None, file_path: str = "") -> bool:
        if not data or not file_path:
            print("⚠️ Missing data or file path")
            return False

        print(f"\n🎨 Generating report from {file_path}...")

        colors = {
            'NOT FOUND':       PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            'MISMATCH':        PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
            'REVIEW REQUIRED': PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid"),
            'VERIFIED MATCH':  PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
        }
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        try:
            if not Path(file_path).exists():
                print(f"❌ File not found: {file_path}")
                return False

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            lc  = ws.max_column
            sc  = lc + 1
            mc  = lc + 2
            rc  = lc + 3
            cc  = lc + 4

            hf = Font(bold=True, size=11)
            for col, title in [
                (sc, "Reconciliation Status"),
                (mc, "Closest KRA Match"),
                (rc, "KRA Reference"),
                (cc, "Confidence Score"),
            ]:
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = hf
                cell.border = thin
                cell.alignment = Alignment(horizontal='center')

            for item in data:
                row_idx = item.get('excel_row')
                if not row_idx or row_idx < 2:
                    continue
                status    = item.get('status', 'UNKNOWN')
                best_m    = item.get('best_match_kra_citation', 'N/A')
                best_r    = item.get('best_match_kra_ref', 'N/A')
                conf      = item.get('confidence_score', '0%')

                ws.cell(row=row_idx, column=sc, value=status).border  = thin
                ws.cell(row=row_idx, column=mc, value=best_m).border  = thin
                ws.cell(row=row_idx, column=rc, value=best_r).border  = thin
                ws.cell(row=row_idx, column=cc, value=conf).border    = thin

                fill = colors.get(status)
                if fill:
                    for col in range(1, cc + 1):
                        ws.cell(row=row_idx, column=col).fill = fill

            for col in ws.columns:
                mx = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 50)

            ts   = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            sd   = Path("reports")
            sd.mkdir(parents=True, exist_ok=True)
            out  = sd / f"{Path(file_path).stem}_RECONCILED_{ts}.xlsx"
            wb.save(out)
            print(f"✅ Report saved to: {out.resolve()}")

            if "temp" in str(file_path) and Path(file_path).exists():
                try: Path(file_path).unlink()
                except: pass

            return True

        except Exception as e:
            errhandler(e, log="report", path="scrapper")
            print(f"❌ Report error: {e}")
            traceback.print_exc()
            return False

    def get_status_summary(self, data=None):
        if data is None:
            data = self.results
        if not data:
            return {}
        total = len(data)
        s = {
            'total':     total,
            'verified':  sum(1 for d in data if d.get('status') == 'VERIFIED MATCH'),
            'review':    sum(1 for d in data if d.get('status') == 'REVIEW REQUIRED'),
            'mismatch':  sum(1 for d in data if d.get('status') == 'MISMATCH'),
            'not_found': sum(1 for d in data if d.get('status') == 'NOT FOUND'),
        }
        if total > 0:
            for k in ['verified','review','mismatch','not_found']:
                s[f'{k}_pct'] = round(s[k] / total * 100, 2)
        return s